import csv
import os
import openpyxl as op
import matplotlib.pyplot as plt

def Parse_Csv_To_List(path):

    with open(path) as file_object:
        reader_file = csv.reader(file_object)
        Raw_Data = list(reader_file)
    return Raw_Data


def Parse_Excel_To_List(path):
    """ 
        Parse a given excel sheet and return list. 
        The data to be parsed should be either in first sheet of excel file or 
        sheet should be named 'Raw Data' (case insensitive).
    """
    WorkBook = op.load_workbook(path)  # WorkBook in which data is available
    DEFAULT_SHEET = "Raw Data"
    try:
        Sheet = WorkBook[DEFAULT_SHEET]  # Particular Sheet in Workbook
    except KeyError as kerr:
        # There is a possiblity that excel file contains 'Raw data' sheet but
        # the letters could be in uppercase or lowercase and cause a mismatch
        DEFAULT_SHEET = DEFAULT_SHEET.lower()
        for sheet in WorkBook.sheetnames:
            if DEFAULT_SHEET == sheet.lower():
                Sheet = sheet
                break
        else:
            # if for-loop doesn't break then it means there was no Raw data sheet in
            # excel file and thus as backup we select the first sheet
            FIRST_SHEET = WorkBook.sheetnames[0]
            Sheet = WorkBook[FIRST_SHEET]
            # Sheet = WorkBook.active               #Last sheet found in excel
    except Exception as e:
        # Unknown/Unexpected error occurred.
        print(f"Exception, at line {e.__traceback__.tb_lineno},",  *e.args)
        print(e.__traceback__.tb_frame, "\n")
        return []

    Total_rows = Sheet.max_row  # Total no. of rows in Sheet
    Total_clms = Sheet.max_column  # Total no. of coloms in Sheet

    RawData = []

    for i in range(1, Total_rows+1):  # Sheet data --->Data(It is 2d list)
        Score = []
        for j in range(1, Total_clms+1):
            if i > 1 and j >= 10:
                Score.append(str(change_data(Sheet.cell(i, j).value)))
            else:
                Score.append(str(Sheet.cell(i, j).value))
        RawData.append(Score)

    WorkBook.close()
    return RawData


def change_data(value):
    d = {
        5: 1,
        4: 2,
        3: 3,
        2: 4,
        1: 5
    }
    return d.get(value)


def value_pi(pi):
    if pi>=16 and pi<=28:
        return "Very Low Stress"
    elif pi>=29 and pi<=41:
        return "Low Stress"
    elif pi>=42 and pi<=54:
        return "Moderate Stress"
    elif pi>=55 and pi<=67:
        return "High Stress"
    elif pi>=68 and pi<=80:
        return "Very High Stress"

def value_ipt(ipt):
    if ipt>=17 and ipt<=30:
        return "Very Low Stress"
    elif ipt>=31 and ipt<=44:
        return "Low Stress"
    elif ipt>=45 and ipt<=57:
        return "Moderate Stress"
    elif ipt>=58 and ipt<=71:
        return "High Stress"
    elif ipt>=72 and ipt<=85:
        return "Very High Stress"
    
def value_fe(fe):
    if fe>=11 and fe<=19:
        return "Very Low Stress"
    elif fe>=20 and fe<=28:
        return "Low Stress"
    elif fe>=29 and fe<=37:
        return "Moderate Stress"
    elif fe>=38 and fe<=46:
        return "High Stress"
    elif fe>=47 and fe<=55:
        return "Very High Stress"
    
def value_ifc(ifc):
    if ifc>=9 and ifc<=16:
        return "Very Low Stress"
    elif ifc>=17 and ifc<=23:
        return "Low Stress"
    elif ifc>=24 and ifc<=30:
        return "Moderate Stress"
    elif ifc>=31 and ifc<=37:
        return "High Stress"
    elif ifc>=38 and ifc<=45:
        return "Very High Stress"
    
def value_peses(peses):
    if peses>=16 and peses<=28:
        return "Very Low Stress"
    elif peses>=29 and peses<=41:
        return "Low Stress"
    elif peses>=42 and peses<=54:
        return "Moderate Stress"
    elif peses>=55 and peses<=67:
        return "High Stress"
    elif peses>=68 and peses<=80:
        return "Very High Stress"

def value_score(sum):
    if sum>=66 and sum<=118:
        return "Very Low Academic Stress"
    elif sum>=119 and sum<=171:
        return "Low Academic Stress"
    elif sum>=172 and sum<=224:
        return "Moderate Academic Stress"
    elif sum>=225 and sum<=277:
        return "High Academic Stress"
    elif sum>=278 and sum<=330:
        return "Very High Stress"

def process(data):
    Total_rows = len(data)  
    Total_cols = len(data[0])

    summary = []
    Raw_Scores = ['PI', 'IPT', 'FE', 'IFC', 'PE&SES']
    Raw_Scores_Value=['Personal Inadequacy (PI)','Interactions with Peers and Teachers (IPT)','Fear of Examination (FE)','Inadequate Facilities at College (IFC)','Parents Expectation and SES (PE&SES)','Total Score','Total Score Interpretation']
    x =y= 0

    for i in range(0, Total_rows):
        Score = []
        pi = ipt = fe = ifc = peses = 0
        a=1
        # Finding values of pi,ipt......
        for b in range(10, Total_cols):
                if i>0 :
                    if b >= 10 and b <= 25:
                        pi += int(data[i][b])
                    elif b > 25 and b <= 42:
                        ipt += int(data[i][b])
                    elif b > 42 and b <= 53:
                        fe += int(data[i][b])
                    elif b > 53 and b <= 62:
                        ifc += int(data[i][b])
                    elif b > 62 and b <= 75:
                        peses += int(data[i][b])
                
        # Adding values of pi,ipt,,, in main data   
        for j in range(0, Total_cols + 6):
            if i == 0 and j > Total_cols:
                Score.append(Raw_Scores[x])
                x += 1
            elif i > 0 and j == 76:
                Score.append(pi)
            elif i > 0 and j == 77:
                Score.append(ipt)
            elif i > 0 and j == 78:
                Score.append(fe)
            elif i > 0 and j == 79:
                Score.append(ifc)
            elif i > 0 and j == 80:
                Score.append(peses)
            elif j<Total_cols:
                Score.append(data[i][j])
            a+=1

        # Adding result of values (pi,ipt,,,) in main data
        for j in range(1,8):
            if i == 0 :
                Score.append(Raw_Scores_Value[y])
                y += 1
            elif j==1:
                Score.append(value_pi(pi))
            elif j==2:
                Score.append(value_ipt(ipt))
            elif j==3:
                Score.append(value_fe(fe))
            elif j==4:
                Score.append(value_ifc(ifc))
            elif j==5:
                Score.append(value_peses(peses))
            elif j==6:
                sum = pi+ipt+fe+ifc+peses
                Score.append(sum)
            elif j==7:
                Score.append(value_score(sum))

        summary.append(Score)
        # pprint(summary)
    return summary

if __name__ == '__main__':
    from pprint import pprint
    FILE = "D:\Projects\Acro care [Stress Reports Generator]\Stress Scale Sample Data.xlsx"
    Rawdata = Parse_Excel_To_List(FILE)
    xyz=process(Rawdata)
    pprint(xyz)
